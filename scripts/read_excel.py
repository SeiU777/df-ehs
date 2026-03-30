"""讀取職災統計 Excel 並輸出為文字檔"""
import sys
import subprocess
import os

try:
    import openpyxl
except ImportError:
    print("正在安裝 openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "公司職災資料", "職災統計__2023.3-2026.3.xlsx")
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "公司職災資料", "職災統計_output.txt")

wb = openpyxl.load_workbook(excel_path)

with open(output_path, "w", encoding="utf-8") as f:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f"\n=== 工作表: {sheet_name} ===\n")
        f.write(f"列數: {ws.max_row}, 欄數: {ws.max_column}\n\n")

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
            vals = []
            for cell in row:
                v = cell.value
                if v is not None:
                    vals.append(str(v))
                else:
                    vals.append("")
            f.write(f"Row{row_idx}: {' | '.join(vals)}\n")

print(f"已輸出至: {output_path}")
print("請回到 Claude 繼續分析。")
